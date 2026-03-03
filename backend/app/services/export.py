"""BigCommerce Excel export: fixed POC columns (5–6), wrap text, no missing data."""
import io
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# POC: fixed columns; Image = image URL; Data Source = Tavily or OpenAI; Source Website = page URL.
BC_COLUMNS = ["Brand Name", "SKU", "Name", "Price", "Description", "Image", "Data Source", "Source Website"]
PLACEHOLDER_IMAGE_URL = "https://via.placeholder.com/300x300?text=No+Image"
PLACEHOLDER_DESCRIPTION = "No description"
PLACEHOLDER_NAME = "Product"


def _safe_get(raw_row: Any, key: str, default: Any = "") -> Any:
    if raw_row is None:
        return default
    if isinstance(raw_row, pd.Series):
        v = raw_row.get(key, default)
        return v if pd.notna(v) else default
    return raw_row.get(key, default)


def _row_to_bc(prod: dict[str, Any], _images_base_path: Path | None) -> dict[str, Any]:
    raw = prod.get("raw_row")
    row = {c: "" for c in BC_COLUMNS}
    row["Brand Name"] = str((prod.get("brand_name") or "").strip() or "—")
    row["SKU"] = str((prod.get("sku_raw") or prod.get("sku") or "").strip() or "—")
    row["Name"] = str((prod.get("name") or "").strip() or (row["SKU"] if row["SKU"] != "—" else PLACEHOLDER_NAME))
    row["Price"] = str(prod.get("price") or _safe_get(raw, "Retail Price") or _safe_get(raw, "List Price") or "0")
    row["Description"] = str((prod.get("description") or "").strip() or PLACEHOLDER_DESCRIPTION)
    # Use same canonical image URL as BC (from chosen source website only)
    canonical_image = (prod.get("image_url") or "").strip()
    if not canonical_image:
        image_urls = prod.get("_image_urls") or []
        canonical_image = (image_urls[0] if image_urls else "").strip()
    row["Image"] = str(canonical_image or PLACEHOLDER_IMAGE_URL)
    method = (prod.get("_search_method") or "").strip().lower()
    row["Data Source"] = "OpenAI" if method == "openai" else ("Tavily" if method == "tavily" else "—")
    row["Source Website"] = str((prod.get("source_website") or "").strip() or "—")
    return row


def build_bc_dataframe(enriched_products: list[dict[str, Any]], images_base_path: Path | None = None) -> pd.DataFrame:
    rows = [_row_to_bc(p, images_base_path) for p in enriched_products]
    return pd.DataFrame(rows, columns=BC_COLUMNS)


def _apply_formatting(wb) -> None:
    ws = wb.active
    if not ws:
        return
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        w = min(max(len(str(ws.cell(row=1, column=col_idx).value or "")) + 2, 10), 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Products") -> io.BytesIO:
    buf = io.BytesIO()
    df.to_excel(buf, index=False, sheet_name=sheet_name)
    buf.seek(0)
    wb = load_workbook(buf)
    _apply_formatting(wb)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_to_excel(enriched_products: list[dict[str, Any]], output_path: str | Path, images_base_path: Path | None = None) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = build_bc_dataframe(enriched_products, images_base_path)
    df.to_excel(output_path, index=False, sheet_name="Products")
    wb = load_workbook(output_path)
    _apply_formatting(wb)
    wb.save(output_path)
    wb.close()
    return output_path
